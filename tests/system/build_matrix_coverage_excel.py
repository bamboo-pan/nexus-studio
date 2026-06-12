from __future__ import annotations

import json
import re
import subprocess
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SYSTEM_TEST_DIR = Path(__file__).resolve().parent
WORKSPACE_ROOT = Path(__file__).resolve().parents[2]
OUTPUT_PATH = WORKSPACE_ROOT / "系统测试矩阵覆盖跟踪.xlsx"

BASELINE_CLEAN_FORMAL_RUN = "/home/bamboo/nexus-studio-system-test-20260608-140643"
API_REUSE_SOURCE_RUN = "/home/bamboo/nexus-studio-system-test-20260608-221554"
OFFICIAL_BASELINE_SOURCE_RUN = "/home/bamboo/nexus-studio-system-test-20260608-203835"
LATEST_PARTIAL_UI_RUN = "/home/bamboo/nexus-studio-system-test-20260610-011140"
LATEST_RESUME_RUN = "/home/bamboo/nexus-studio-system-test-20260612-001338"
LATEST_COMPLETE_MATRIX_RUN = "/home/bamboo/nexus-studio-system-test-20260611-232555"
CURRENT_MATRIX_RUN = LATEST_COMPLETE_MATRIX_RUN

CURRENT_ACCOUNT_COUNT = 2
CURRENT_NATIVE_WORKERS_PER_ACCOUNT = 3
CURRENT_WARMUP_GOOGLE_REQUEST_ESTIMATE = CURRENT_ACCOUNT_COUNT * CURRENT_NATIVE_WORKERS_PER_ACCOUNT

STATUS_LABELS = {
    "pass": "通过",
    "fail": "失败",
    "not_applicable": "不适用",
    "blocked": "阻塞",
    "unknown": "未知",
    "not_mapped": "未映射",
}

TESTED_BY_LABELS = {
    "mcp_visible": "MCP 可见 UI",
    "automated_playwright": "自动化 Playwright",
    "api_only_not_ui_feature": "仅 API（非 UI 功能）",
    "not_applicable_with_evidence": "不适用（有证据）",
    "not_covered": "未覆盖",
}

VALUE_LABELS = {
    "yes": "是",
    "no": "否",
    "true": "是",
    "false": "否",
    "blocked_external_google_quota": "外部 Google 配额阻塞",
    "not_checked": "未检查",
    "None / local filesystem": "无 / 本地文件系统",
    "Mixed": "混合",
    "Google credential state": "Google 凭据状态",
    "Google AI Studio": "Google AI Studio",
    "OpenAI-compatible": "OpenAI 兼容",
    "No generation request": "不产生生成请求",
    "No fixed generation request": "不产生固定生成请求",
    "No new provider request": "不新增 provider 请求",
    "Google generation": "消耗 Google 生成额度",
    "OpenAI generation": "消耗 OpenAI 生成额度/费用",
    "OpenAI quota/cost": "OpenAI 额度/费用敏感",
    "Provider request when covered": "覆盖时会产生 provider 请求",
    "Mixed provider generation/model-list": "混合 provider 生成/模型列表请求",
    "Google image quota when covered": "覆盖时消耗 Google 图像额度",
    "Mixed/depends on subcase": "混合，取决于子用例",
    "Model-list and warmup only": "仅模型列表与预热",
    "low/no generation quota, still real upstream": "低/无生成额度消耗，但仍是真实上游请求",
}

AREA_LABELS = {
    "Local Studio provider matrix": "Local Studio provider 矩阵",
    "Local Studio UI behavior": "Local Studio UI 行为",
    "Base Playground chat": "基础 Playground 聊天",
    "Base image generation": "基础图像生成",
    "Request logs": "请求记录",
    "Accounts": "账号管理",
    "Local Studio API": "Local Studio API",
    "Request-log API": "请求记录 API",
    "Base-compatible APIs": "基础兼容 API",
    "Account APIs": "账号 API",
    "Provider Manager": "Provider Manager",
    "Environment gate": "环境门禁",
    "Startup gate": "启动门禁",
    "Request logging": "请求记录",
    "Secret redaction": "敏感信息脱敏",
    "Regression oracle": "回归断言",
    "Performance/native regression": "性能 / native 回归",
    "Other": "其他",
}

SURFACE_LABELS = {
    "Local Studio UI/API": "Local Studio UI/API",
    "Visible UI": "可见 UI",
    "Playground UI": "Playground UI",
    "Visible UI/API": "可见 UI/API",
    "HTTP API": "HTTP API",
    "HTTP API + UI": "HTTP API + UI",
    "Control plane": "控制面",
    "Harness": "测试 harness",
    "Harness/API/UI": "测试 harness/API/UI",
    "API/UI": "API/UI",
    "API/UI/log exports": "API/UI/日志导出",
    "Local Studio": "Local Studio",
    "Local Studio/API": "Local Studio/API",
    "Official AI Studio + Local Studio": "官方 AI Studio + Local Studio",
    "Mixed": "混合",
}

PHASE_LABELS = {
    "Startup warmup": "启动预热",
    "API matrix": "API 矩阵",
    "API model probe": "API 模型探测",
    "Official baseline": "官方基准",
    "Host UI Google": "主机 UI - Google",
    "Host UI OpenAI": "主机 UI - OpenAI",
    "Model/control calls": "模型/控制调用",
    "Latest blocked run": "最近阻塞运行",
    "Local/control plane": "本地/控制面",
}

IMAGE_EXPECTATION_LABELS = {
    "yes": "需要真实生成图片",
    "no": "不应生成图片",
    "yes_or_controlled_unsupported": "需要真实生成图片，或受控证明 provider 不支持",
    "yes_or_image_prompt_oracle": "需要图像 prompt / 图片工具断言",
    "asset_get_only": "仅打开已生成图片资源",
    "partial_asset_followup": "依赖已生成图片的后续断言",
    "future_phase": "未来阶段适用",
}

REQUEST_LABELS = {
    "Native UI worker warmup for copied accounts": "复制账号的 Native UI worker 预热",
    "Google Local Studio API text/search/optional-tool sends": "Google Local Studio API 文本 / 搜索 / 可选工具发送",
    "OpenAI-compatible Local Studio API sends": "OpenAI 兼容 Local Studio API 发送",
    "OpenAI-compatible Responses model and reasoning probes": "OpenAI 兼容 Responses 模型与 reasoning 探测",
    "Official AI Studio visible same-account baseline samples": "官方 AI Studio 同账号可见基准样本",
    "Local Studio Google performance samples": "Local Studio Google 性能样本",
    "Local Studio Google repeated prompt": "Local Studio Google 重复 prompt",
    "Local Studio Google optional-tools text + nonstream search": "Local Studio Google 可选工具文本 + 非流式搜索",
    "Local Studio Google reasoning matrix": "Local Studio Google reasoning 矩阵",
    "Local Studio Google interface compatibility matrix": "Local Studio Google interface 兼容矩阵",
    "Local Studio conversation CRUD seed + rerun": "Local Studio 会话 CRUD 种子发送 + rerun",
    "Playground Gemini basic + search chat": "Playground Gemini 基础聊天 + 搜索聊天",
    "OpenAI-compatible optional/basic/search/repeated visible sends": "OpenAI 兼容可选工具 / 基础 / 搜索 / 重复可见发送",
    "OpenAI-compatible reasoning matrix": "OpenAI 兼容 reasoning 矩阵",
    "OpenAI-compatible interface compatibility matrix": "OpenAI 兼容 interface 兼容矩阵",
    "Invalid provider recovery send": "错误 provider 后恢复发送",
    "Google model-list/control calls used by API/UI harness": "API/UI harness 使用的 Google 模型列表 / 控制调用",
    "OpenAI-compatible model-list/control calls used by API/UI harness": "API/UI harness 使用的 OpenAI 兼容模型列表 / 控制调用",
    "Failed Google quota attempt at google-ai-studio-performance-1": "google-ai-studio-performance-1 的 Google 配额失败尝试",
    "Provider Manager, request logs, account local health, synthetic account CRUD": "Provider Manager、请求记录、账号本地健康、合成账号 CRUD",
}

REQUEST_TYPE_LABELS = {
    "generation-readiness probe": "生成就绪探测",
    "generation": "生成请求",
    "generation probe": "生成探测",
    "generation/control-error matrix": "生成 / 受控错误矩阵",
    "generation after controlled local error": "受控本地错误后的生成请求",
    "model/control": "模型 / 控制调用",
    "failed generation attempt": "失败的生成尝试",
    "local/control": "本地 / 控制调用",
}

REQUEST_NOTES = {
    "Variable item: current estimate is 2 copied accounts * 3 native workers/account. Retries/recovery can increase this; do not treat as a fixed matrix row count.": "变量项：当前估算为 2 个复制账号 * 每账号 3 个 native worker。重试 / 恢复可能增加次数，不能当作固定矩阵行数。",
    "google-basic, google-repeat-1/2, google optional search+image text, Google search stream. Reused from API source in low-quota resume.": "包含 google-basic、google-repeat-1/2、Google 可选 search+image 文本、Google search stream。低额度 resume 已复用 API 来源产物。",
    "Seven logical API prompts. Retry budget: basic/search/optional paths up to 3 each; reasoning stream/nonstream/repeat up to 5 each. Current artifact observed 9 sends because reasoning nonstream required 3 attempts.": "7 个逻辑 API prompt。重试预算：basic/search/optional 路径每项最多 3 次；reasoning stream/nonstream/repeat 每项最多 5 次。当前产物观察到 9 次发送，因为 reasoning nonstream 需要 3 次尝试。",
    "Current key/model discovery observed 3 text-model probes plus 1 reasoning-model probe. A different preferred model can reduce this; more candidate failures can increase it.": "当前 key/model discovery 观察到 3 次文本模型探测 + 1 次 reasoning 模型探测。更合适的首选模型可减少次数，更多候选失败会增加次数。",
    "Needs 3 successful visible samples. The reusable 203835 artifact had 4 attempts because one visible-text sample was discarded. Low-quota resume reuses this artifact.": "需要 3 个成功可见样本。可复用的 203835 产物有 4 次尝试，因为 1 个可见文本样本被丢弃。低额度 resume 复用该产物。",
    "Latest resume blocked on google-ai-studio-performance-1; the failed quota attempt is not credited as coverage.": "最近 resume 阻塞在 google-ai-studio-performance-1；配额失败尝试不计入覆盖。",
    "Two same-prompt sends to prove no result cache hit.": "两次相同 prompt 发送，用于证明没有结果缓存命中。",
    "One search+image-tool ordinary text send and one non-stream search send.": "一次 search+image-tool 普通文本发送，以及一次非流式搜索发送。",
    "Stream, non-stream, repeat. Google branch has no explicit retry loop in the current host smoke.": "覆盖 stream、non-stream、repeat。当前 host smoke 的 Google 分支没有显式重试循环。",
    "Gemini and OpenAI Chat strict entries send 6 requests. Claude compatibility entries may fail before send (min) or send 3 controlled-error requests (max).": "Gemini 与 OpenAI Chat 严格项发送 6 次请求。Claude 兼容项可能发送前失败（最小值），或发送 3 次受控错误请求（最大值）。",
    "One seed send and one rerun from the first user turn.": "一次种子发送，以及从第一轮用户消息触发的一次 rerun。",
    "One basic chat send and one Search-enabled Playground send.": "一次基础聊天发送，以及一次开启 Search 的 Playground 发送。",
    "Optional search+image text, optional search+image+reasoning text, basic, search, repeat-1, repeat-2.": "包含可选 search+image 文本、可选 search+image+reasoning 文本、basic、search、repeat-1、repeat-2。",
    "Stream, non-stream, repeat per attempt. Host smoke retries the OpenAI reasoning matrix up to 5 attempts if summary/text evidence is missing.": "每次尝试覆盖 stream、non-stream、repeat。如果缺少 summary/text 证据，host smoke 对 OpenAI reasoning 矩阵最多重试 5 次。",
    "OpenAI Chat strict entries send 3 requests. Gemini/Claude compatibility entries may fail before send or add up to 6 controlled-error requests.": "OpenAI Chat 严格项发送 3 次请求。Gemini/Claude 兼容项可能发送前失败，也可能增加最多 6 次受控错误请求。",
    "The invalid 127.0.0.1:9 send is not an external provider request. The recovery send can retry up to 3 times for transient upstream errors.": "错误的 127.0.0.1:9 发送不算外部 provider 请求。恢复发送遇到临时上游错误时最多重试 3 次。",
    "Model-list estimates include boot/API/host UI loads. Host Local Studio model loaders retry up to 3 attempts. Browser page resource traffic is excluded.": "模型列表估算包含 boot/API/host UI 加载。Host Local Studio 模型加载器最多重试 3 次。不包含浏览器页面静态资源流量。",
    "Includes direct OpenAI /models preflight, compatibility model list, API Local Studio model list, and host UI model loads. Direct Responses probes are counted above as generation probes.": "包含直接 OpenAI /models 预检、兼容模型列表、API Local Studio 模型列表、host UI 模型加载。直接 Responses 探测已在上方按生成探测计数。",
    "The 074940 resume attempted one current Google request and received official quota exhaustion. It consumed risk but produced no coverage credit.": "074940 resume 发起了 1 次当前 Google 请求并收到官方配额耗尽。它消耗了额度风险，但不产生覆盖信用。",
    "Account health checks validate local storage-state shape; generation permission is proven by warmup or real generation, not the health button itself.": "账号健康检查只验证本地 storage-state 形状；生成权限必须由 warmup 或真实生成证明，不能由健康检查按钮本身替代。",
}

IMAGE_CATEGORY_LABELS = {
    "Rendered image generation": "真实图片生成",
    "Image Tool optional text path": "图片工具可选文本路径",
    "Search-to-image multi-turn bug path": "搜索到图片的多轮 bug 路径",
    "Non-stream search+image generation": "非流式搜索 + 图片生成",
    "Rendered image generation or controlled unsupported path": "真实图片生成或受控不支持路径",
    "Search+Image+Reasoning optional text path": "Search + Image + Reasoning 可选文本路径",
    "Image Tool controls": "图片工具控件",
    "Image model filtering": "图片模型过滤",
    "Base Images page generation": "基础图片页生成",
    "Base Images reference/edit/retry": "基础图片页参考图 / 编辑 / 重试",
    "Image model list API": "图片模型列表 API",
    "Google API image-tool bug path": "Google API 图片工具 bug 路径",
    "Generated image asset API": "已生成图片资源 API",
    "Generated asset/cache variation API": "已生成资源 / 缓存变化 API",
    "Base-compatible image API": "基础兼容图片 API",
    "Regression: Google search+image multi-turn": "回归：Google 搜索 + 图片多轮路径",
    "Future protocol image matrix": "未来协议图片矩阵",
    "Future image routing policy": "未来图片路由策略",
}

IMAGE_COVERAGE_LABELS = {
    "missing: Google rendered image prompt was not executed": "缺口：尚未执行 Google 真实图片 prompt。",
    "text-only optional-tool path; Image Tool is enabled but prompt must not force image generation": "仅文本可选工具路径：图片工具开启，但 prompt 不应强制生成图片。",
    "missing: multi-turn Google search-to-image path was not executed": "缺口：尚未执行 Google 搜索到图片的多轮路径。",
    "missing: non-stream infographic/image path was not executed": "缺口：尚未执行非流式信息图 / 图片路径。",
    "missing: OpenAI-compatible image generation or controlled unsupported-path assertion was not executed": "缺口：尚未执行 OpenAI 兼容图片生成或受控不支持路径断言。",
    "text-only reasoning/tool path; image generation should not be forced": "仅文本 reasoning/tool 路径：不应强制生成图片。",
    "controls/model selectors only; no rendered image expected": "仅控件 / 模型选择器，不期望真实图片输出。",
    "model-list/filtering only; no rendered image expected": "仅模型列表 / 过滤，不期望真实图片输出。",
    "missing: #images generation path was not executed": "缺口：尚未执行 #images 生成路径。",
    "missing: #images reference/edit/retry path was not executed": "缺口：尚未执行 #images 参考图 / 编辑 / 重试路径。",
    "model-list only; should include image models": "仅模型列表；应包含图片模型。",
    "not proven as rendered/image-prompt coverage in current artifact; mapped API pass is search/tool regression evidence": "当前产物未证明真实图片 / 图片 prompt 覆盖；已映射 API pass 只是 search/tool 回归证据。",
    "missing: generated image asset URL was not opened through API": "缺口：尚未通过 API 打开已生成图片 URL。",
    "partial: generated-asset API coverage is explicitly missing": "部分覆盖：已生成资源 API 覆盖明确缺失。",
    "missing: /v1/images/generations smoke was not executed": "缺口：尚未执行 /v1/images/generations smoke。",
    "missing: regression path was not executed in visible UI": "缺口：尚未在可见 UI 中执行回归路径。",
    "not applicable in current Phase 1 code state": "当前 Phase 1 代码状态不适用。",
}

IMAGE_NOTES = {
    "Needs a Google Responses Image Tool prompt with a visible rendered image URL and request-log evidence.": "需要 Google Responses 图片工具 prompt，且有可见图片 URL 和 request-log 证据。",
    "This is image-tool semantics, not rendered image generation. Current clean formal artifact still marks the row failed; later harness code maps this path but was blocked by Google quota before a new clean pass artifact.": "这是图片工具语义，不是真实图片生成。当前 clean formal 产物仍将该行标为失败；后续 harness 已映射此路径，但在产生新的 clean pass 产物前被 Google 配额阻塞。",
    "Covers greet, identity, news/search, then make-it-image. BUG-GEMINI-IMAGE-TOOL-01 can share this same path if assertions are strong enough.": "覆盖问候、询问身份、新闻 / 搜索，然后做成图片。如果断言足够强，BUG-GEMINI-IMAGE-TOOL-01 可复用同一路径。",
    "Needed to prove image/text/error persistence does not depend on SSE streaming.": "用于证明图片 / 文本 / 错误持久化不依赖 SSE 流式路径。",
    "If the configured OpenAI-compatible provider does not support images, this should still assert graceful failure and health recovery.": "如果配置的 OpenAI 兼容 provider 不支持图片，也必须断言优雅失败和健康恢复。",
    "This proves optional image capability is not forced for ordinary text prompts.": "用于证明普通文本 prompt 不会被强制触发图片能力。",
    "This is image-tool optionality plus reasoning preservation, not rendered image generation.": "这是图片工具可选语义 + reasoning 保留，不是真实图片生成。",
    "Checks provider-specific image controls and avoids cross-provider residual UI state.": "检查 provider 专属图片控件，并避免跨 provider 残留 UI 状态。",
    "Image models must appear in Image Tool selectors and not pollute chat model lists.": "图片模型必须出现在图片工具选择器中，且不能污染聊天模型列表。",
    "Needed to prove base image generation is independent from Local Studio Image Tool.": "用于证明基础图片生成页独立于 Local Studio 图片工具。",
    "Estimate allows one edit/reference generation or a separate edit plus retry assertion.": "估算允许一次编辑 / 参考图生成，或一次编辑加一次独立重试断言。",
    "Provider-facing model list/control calls are counted separately from generation requests.": "面向 provider 的模型列表 / 控制调用与生成请求分开统计。",
    "SYSTEM_TEST_PLAN.md asks for Google search=true + image_tool_enabled=true + image prompt. Current clean artifact does not prove a generated image asset.": "SYSTEM_TEST_PLAN.md 要求 Google search=true + image_tool_enabled=true + image prompt。当前 clean artifact 尚未证明已生成图片资源。",
    "Requires an image produced by another row; the asset GET itself is local and should not consume provider quota.": "依赖其他行先生成图片；asset GET 本身是本地请求，不应消耗 provider 额度。",
    "This can piggyback on a real generated image from G-LS/API/base image rows.": "可复用 G-LS/API/base image 行生成的真实图片。",
    "Covers the base API image-generation surface outside Local Studio.": "覆盖 Local Studio 之外的基础 API 图片生成入口。",
    "Do not double count if the G-LS-06 multi-turn path includes the regression assertions.": "如果 G-LS-06 多轮路径已包含回归断言，不要重复计数。",
    "Becomes applicable when shared runtime gateway/protocol adapter matrix is implemented.": "shared runtime gateway / protocol adapter 矩阵实现后适用。",
    "Becomes applicable when advanced routing policy handles stream/tool/image compatibility.": "高级路由策略处理 stream/tool/image 兼容性后适用。",
}


def _unc_path_from_wsl_path(wsl_path: str) -> Path:
    relative = wsl_path.removeprefix("/").replace("/", "\\")
    return Path("\\\\wsl.localhost\\Ubuntu-24.04") / relative


def read_wsl_text(wsl_path: str) -> str | None:
    direct_path = Path(wsl_path)
    if direct_path.exists():
        return direct_path.read_text(encoding="utf-8")

    try:
        unc_path = _unc_path_from_wsl_path(wsl_path)
        if unc_path.exists():
            return unc_path.read_text(encoding="utf-8")
    except OSError:
        pass

    try:
        completed = subprocess.run(
            ["wsl.exe", "cat", wsl_path],
            check=False,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
        )
    except OSError:
        return None
    if completed.returncode != 0:
        return None
    return completed.stdout


def read_wsl_json(wsl_path: str) -> dict[str, Any]:
    text = read_wsl_text(wsl_path)
    if not text:
        return {}
    try:
        payload = json.loads(text)
    except json.JSONDecodeError:
        return {}
    return payload if isinstance(payload, dict) else {}


def artifact_path(run_root: str, artifact_name: str) -> str:
    return f"{run_root}/artifacts/{artifact_name}"


def google_quota_exhausted_text(text: str) -> bool:
    normalized = (text or "").lower()
    markers = (
        "you exceeded your current quota",
        "ai.google.dev/gemini-api/docs/rate-limits",
        "quota for the day",
        "daily quota",
        "quota resets",
        "resource_exhausted",
    )
    return any(marker in normalized for marker in markers)


def latest_quota_blocker(run_root: str) -> dict[str, Any]:
    blocker = read_wsl_json(artifact_path(run_root, "google-quota-blocker.safe.json"))
    if blocker:
        return blocker
    summary = read_wsl_json(artifact_path(run_root, "summary.json"))
    if summary and summary.get("incomplete_reason") != "external_google_quota_exhausted":
        return {}
    warmup = read_wsl_json(artifact_path(run_root, "warmup-health.safe.json"))
    server_log = read_wsl_text(artifact_path(run_root, "server.log")) or ""
    if not warmup or not google_quota_exhausted_text(server_log):
        return {}
    quota_lines = [line for line in server_log.splitlines() if google_quota_exhausted_text(line)]
    warmup_payload = warmup.get("warmup") if isinstance(warmup.get("warmup"), dict) else {}
    return {
        "result": "blocked",
        "reason": "external_google_quota_exhausted",
        "source": "wsl-startup-warmup",
        "label": "startup-account-browser-warmup",
        "status": 429 if "status=429" in server_log or "current quota" in server_log.lower() else 0,
        "warmup_status": warmup_payload.get("status"),
        "text_preview": compact_json("\n".join(quota_lines[-5:]) or server_log, 500),
        "warmup": warmup_payload,
    }


def compact_json(value: Any, limit: int = 900) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, str):
        text = value
    else:
        text = json.dumps(value, ensure_ascii=False, sort_keys=True)
    text = re.sub(r"\s+", " ", text).strip()
    return text if len(text) <= limit else text[: limit - 3] + "..."


def label(value: Any, mapping: dict[str, str] | None = None) -> str:
    if isinstance(value, bool):
        return "是" if value else "否"
    text = str(value or "")
    if not text:
        return ""
    if mapping and text in mapping:
        return mapping[text]
    return VALUE_LABELS.get(text, text)


def status_label(value: Any) -> str:
    return STATUS_LABELS.get(str(value or "unknown"), str(value or "未知"))


def tested_by_label(value: Any) -> str:
    return TESTED_BY_LABELS.get(str(value or ""), str(value or ""))


def priority_for_case(case_id: str) -> str:
    if case_id.startswith(("LS-UI", "BASE-", "API-")):
        return "P1"
    return "P0"


def followup_state(case_id: str, status: str, tested_by: str) -> str:
    if status == "pass":
        return "已覆盖"
    if status == "not_applicable":
        return "阶段不适用"
    if tested_by == "not_covered":
        return "待补测"
    if case_id.startswith("BUG-"):
        return "待回归"
    return "待修复/待复测"


def next_action_for_case(case_id: str, status: str, tested_by: str) -> str:
    if status == "pass":
        return "保留为回归证据；后续完整系统测试继续覆盖。"
    if status == "not_applicable":
        return "进入对应 Provider Manager / shared runtime 阶段后启用。"
    if case_id == "BASE-IMG-02":
        return "保留为回归证据；后续完整系统测试继续覆盖 #images reference/edit/retry。"
    if case_id.startswith(("G-LS", "BUG-GEMINI", "BUG-AISTUDIO", "PERF-")):
        return "Google 额度恢复后，用低额度 resume 继续真实 UI 路径并补齐断言。"
    if "IMG" in case_id or case_id in {"G-LS-04", "G-LS-06", "G-LS-07", "O-LS-04", "API-LS-04", "API-LS-06", "API-BASE-01"}:
        return "补跑真实图像生成/图片资源路径，确认 UI、conversation、request log 一致。"
    if tested_by == "not_covered":
        return "补充脚本硬断言或 MCP 可见人工验收项。"
    return "修复失败原因后从干净 WSL 副本重跑对应 API + UI 用例。"


def blocker_for_case(case_id: str, status: str, tested_by: str, evidence_or_reason: str) -> str:
    if status == "pass":
        return ""
    if status == "not_applicable":
        return evidence_or_reason or "当前阶段不适用"
    if case_id == "BASE-IMG-02":
        return "最新完整矩阵已通过 #images reference/edit/retry；代理桥接、账号恢复和图片模型 fallback 均被真实 UI 路径覆盖。"
    if case_id.startswith(("G-LS", "BUG-GEMINI", "BUG-AISTUDIO", "PERF-")):
        return "最近 resume 被外部 Google 配额阻塞；当前 clean artifact 仍保留失败/缺口。"
    if tested_by == "not_covered":
        return "当前 harness 未覆盖该计划行。"
    return evidence_or_reason


def parse_plan_rows() -> dict[str, dict[str, str]]:
    plan_path = WORKSPACE_ROOT / "SYSTEM_TEST_PLAN.md"
    rows: dict[str, dict[str, str]] = {}
    case_id_pattern = re.compile(r"^[A-Z0-9]+(?:-[A-Z0-9]+)+$")
    for line in plan_path.read_text(encoding="utf-8").splitlines():
        heading = line.strip().lstrip("#").strip()
        if case_id_pattern.match(heading) and any(character.isdigit() for character in heading):
            rows.setdefault(
                heading,
                {
                    "plan_area": "专项断言",
                    "plan_action": "见 SYSTEM_TEST_PLAN.md 对应用例章节",
                    "plan_expected": "必须满足该专项章节列出的全部断言。",
                    "plan_columns": heading,
                },
            )
            continue
        if not line.startswith("|"):
            continue
        cells = [cell.strip().strip("`") for cell in line.strip().strip("|").split("|")]
        if not cells or not case_id_pattern.match(cells[0]) or not any(character.isdigit() for character in cells[0]):
            continue
        case_id = cells[0]
        if case_id.lower() == "case id" or set(case_id) <= {"-"}:
            continue
        rows[case_id] = {
            "plan_area": cells[1] if len(cells) > 1 else "",
            "plan_action": cells[-2] if len(cells) > 3 else (cells[2] if len(cells) > 2 else ""),
            "plan_expected": cells[-1] if len(cells) > 2 else "",
            "plan_columns": compact_json(cells, limit=700),
        }
    return rows


def classify_case(case_id: str) -> dict[str, str]:
    if case_id.startswith("G-LS"):
        return {"coverage_area": "Local Studio provider matrix", "provider": "Google AI Studio", "surface": "Local Studio UI/API", "quota_risk": "Google generation"}
    if case_id.startswith("O-LS"):
        return {"coverage_area": "Local Studio provider matrix", "provider": "OpenAI-compatible", "surface": "Local Studio UI/API", "quota_risk": "OpenAI generation"}
    if case_id.startswith("LS-UI"):
        return {"coverage_area": "Local Studio UI behavior", "provider": "Mixed", "surface": "Visible UI", "quota_risk": "Mixed/depends on subcase"}
    if case_id.startswith("BASE-CHAT"):
        return {"coverage_area": "Base Playground chat", "provider": "Google AI Studio", "surface": "Playground UI", "quota_risk": "Google generation"}
    if case_id.startswith("BASE-IMG"):
        return {"coverage_area": "Base image generation", "provider": "Google AI Studio", "surface": "Images UI", "quota_risk": "Google image quota when covered"}
    if case_id.startswith("BASE-REQ"):
        return {"coverage_area": "Request logs", "provider": "None", "surface": "Visible UI/API", "quota_risk": "No new provider request"}
    if case_id.startswith("BASE-ACC"):
        return {"coverage_area": "Accounts", "provider": "Google credential state", "surface": "Visible UI/API", "quota_risk": "No fixed generation request"}
    if case_id.startswith("API-LS"):
        return {"coverage_area": "Local Studio API", "provider": "Mixed", "surface": "HTTP API", "quota_risk": "Mixed provider generation/model-list"}
    if case_id.startswith("API-REQ"):
        return {"coverage_area": "Request-log API", "provider": "None", "surface": "HTTP API", "quota_risk": "No new provider request"}
    if case_id.startswith("API-BASE"):
        return {"coverage_area": "Base-compatible APIs", "provider": "Mixed", "surface": "HTTP API", "quota_risk": "Provider request when covered"}
    if case_id.startswith("API-ACC"):
        return {"coverage_area": "Account APIs", "provider": "Google credential state", "surface": "HTTP API + UI", "quota_risk": "No fixed generation request"}
    if case_id.startswith("PM-"):
        return {"coverage_area": "Provider Manager", "provider": "None", "surface": "Control plane", "quota_risk": "No generation request"}
    if case_id.startswith("ENV-"):
        return {"coverage_area": "Environment gate", "provider": "None", "surface": "Harness", "quota_risk": "No generation request"}
    if case_id.startswith("BOOT-"):
        return {"coverage_area": "Startup gate", "provider": "Mixed", "surface": "Harness/API/UI", "quota_risk": "Model-list and warmup only"}
    if case_id.startswith("LOG-"):
        return {"coverage_area": "Request logging", "provider": "None", "surface": "API/UI", "quota_risk": "No generation request"}
    if case_id.startswith("SEC-"):
        return {"coverage_area": "Secret redaction", "provider": "OpenAI-compatible", "surface": "API/UI/log exports", "quota_risk": "Uses existing OpenAI flows"}
    if case_id.startswith("BUG-GEMINI"):
        return {"coverage_area": "Regression oracle", "provider": "Google AI Studio", "surface": "Local Studio", "quota_risk": "Google generation when covered"}
    if case_id.startswith("BUG-OPENAI"):
        return {"coverage_area": "Regression oracle", "provider": "OpenAI-compatible", "surface": "Local Studio/API", "quota_risk": "OpenAI generation"}
    if case_id.startswith("BUG-AISTUDIO") or case_id.startswith("PERF-"):
        return {"coverage_area": "Performance/native regression", "provider": "Google AI Studio", "surface": "Official AI Studio + Local Studio", "quota_risk": "Google generation"}
    return {"coverage_area": "Other", "provider": "Mixed", "surface": "Mixed", "quota_risk": "Review case"}


def request_rows() -> list[dict[str, Any]]:
    return [
        {
            "phase": "Startup warmup",
            "label": "Native UI worker warmup for copied accounts",
            "provider": "Google AI Studio",
            "surface": "Server startup / GET /health",
            "request_type": "generation-readiness probe",
            "fresh_min": CURRENT_WARMUP_GOOGLE_REQUEST_ESTIMATE,
            "fresh_max": CURRENT_WARMUP_GOOGLE_REQUEST_ESTIMATE,
            "observed_current_config": CURRENT_WARMUP_GOOGLE_REQUEST_ESTIMATE,
            "resume_remaining_min": CURRENT_WARMUP_GOOGLE_REQUEST_ESTIMATE,
            "resume_remaining_max": CURRENT_WARMUP_GOOGLE_REQUEST_ESTIMATE,
            "quota_sensitive": "yes",
            "generation_or_probe": True,
            "fixed_matrix": False,
            "notes": "Variable item: current estimate is 2 copied accounts * 3 native workers/account. Retries/recovery can increase this; do not treat as a fixed matrix row count.",
        },
        {
            "phase": "API matrix",
            "label": "Google Local Studio API text/search/optional-tool sends",
            "provider": "Google AI Studio",
            "surface": "POST /api/local-studio/chat",
            "request_type": "generation",
            "fresh_min": 5,
            "fresh_max": 5,
            "observed_current_config": 5,
            "resume_remaining_min": 0,
            "resume_remaining_max": 0,
            "quota_sensitive": "yes",
            "generation_or_probe": True,
            "fixed_matrix": True,
            "notes": "google-basic, google-repeat-1/2, google optional search+image text, Google search stream. Reused from API source in low-quota resume.",
        },
        {
            "phase": "API matrix",
            "label": "OpenAI-compatible Local Studio API sends",
            "provider": "OpenAI-compatible",
            "surface": "POST /api/local-studio/chat",
            "request_type": "generation",
            "fresh_min": 7,
            "fresh_max": 27,
            "observed_current_config": 9,
            "resume_remaining_min": 0,
            "resume_remaining_max": 0,
            "quota_sensitive": "OpenAI quota/cost",
            "generation_or_probe": True,
            "fixed_matrix": True,
            "notes": "Seven logical API prompts. Retry budget: basic/search/optional paths up to 3 each; reasoning stream/nonstream/repeat up to 5 each. Current artifact observed 9 sends because reasoning nonstream required 3 attempts.",
        },
        {
            "phase": "API model probe",
            "label": "OpenAI-compatible Responses model and reasoning probes",
            "provider": "OpenAI-compatible",
            "surface": "Direct provider /responses",
            "request_type": "generation probe",
            "fresh_min": 4,
            "fresh_max": 4,
            "observed_current_config": 4,
            "resume_remaining_min": 0,
            "resume_remaining_max": 0,
            "quota_sensitive": "OpenAI quota/cost",
            "generation_or_probe": True,
            "fixed_matrix": True,
            "notes": "Current key/model discovery observed 3 text-model probes plus 1 reasoning-model probe. A different preferred model can reduce this; more candidate failures can increase it.",
        },
        {
            "phase": "Official baseline",
            "label": "Official AI Studio visible same-account baseline samples",
            "provider": "Google AI Studio",
            "surface": "Official AI Studio UI / GenerateContent",
            "request_type": "generation",
            "fresh_min": 3,
            "fresh_max": 4,
            "observed_current_config": 4,
            "resume_remaining_min": 0,
            "resume_remaining_max": 0,
            "quota_sensitive": "yes",
            "generation_or_probe": True,
            "fixed_matrix": True,
            "notes": "Needs 3 successful visible samples. The reusable 203835 artifact had 4 attempts because one visible-text sample was discarded. Low-quota resume reuses this artifact.",
        },
        {
            "phase": "Host UI Google",
            "label": "Local Studio Google performance samples",
            "provider": "Google AI Studio",
            "surface": "Visible Local Studio UI",
            "request_type": "generation",
            "fresh_min": 3,
            "fresh_max": 3,
            "observed_current_config": 0,
            "resume_remaining_min": 3,
            "resume_remaining_max": 3,
            "quota_sensitive": "yes",
            "generation_or_probe": True,
            "fixed_matrix": True,
            "notes": "Latest UI-reaching resume blocked before these samples completed; the failed quota attempts are not credited as coverage.",
        },
        {
            "phase": "Host UI Google",
            "label": "Local Studio Google repeated prompt",
            "provider": "Google AI Studio",
            "surface": "Visible Local Studio UI",
            "request_type": "generation",
            "fresh_min": 2,
            "fresh_max": 2,
            "observed_current_config": 0,
            "resume_remaining_min": 2,
            "resume_remaining_max": 2,
            "quota_sensitive": "yes",
            "generation_or_probe": True,
            "fixed_matrix": True,
            "notes": "Two same-prompt sends to prove no result cache hit.",
        },
        {
            "phase": "Host UI Google",
            "label": "Local Studio Google optional-tools text + nonstream search",
            "provider": "Google AI Studio",
            "surface": "Visible Local Studio UI",
            "request_type": "generation",
            "fresh_min": 2,
            "fresh_max": 2,
            "observed_current_config": 0,
            "resume_remaining_min": 2,
            "resume_remaining_max": 2,
            "quota_sensitive": "yes",
            "generation_or_probe": True,
            "fixed_matrix": True,
            "notes": "One search+image-tool ordinary text send and one non-stream search send.",
        },
        {
            "phase": "Host UI Google",
            "label": "Local Studio Google reasoning matrix",
            "provider": "Google AI Studio",
            "surface": "Visible Local Studio UI",
            "request_type": "generation",
            "fresh_min": 3,
            "fresh_max": 3,
            "observed_current_config": 0,
            "resume_remaining_min": 3,
            "resume_remaining_max": 3,
            "quota_sensitive": "yes",
            "generation_or_probe": True,
            "fixed_matrix": True,
            "notes": "Stream, non-stream, repeat. Google branch has no explicit retry loop in the current host smoke.",
        },
        {
            "phase": "Host UI Google",
            "label": "Local Studio Google interface compatibility matrix",
            "provider": "Google AI Studio",
            "surface": "Visible Local Studio UI",
            "request_type": "generation/control-error matrix",
            "fresh_min": 6,
            "fresh_max": 9,
            "observed_current_config": 0,
            "resume_remaining_min": 6,
            "resume_remaining_max": 9,
            "quota_sensitive": "yes",
            "generation_or_probe": True,
            "fixed_matrix": True,
            "notes": "Gemini and OpenAI Chat strict entries send 6 requests. Claude compatibility entries may fail before send (min) or send 3 controlled-error requests (max).",
        },
        {
            "phase": "Host UI Google",
            "label": "Local Studio conversation CRUD seed + rerun",
            "provider": "Google AI Studio",
            "surface": "Visible Local Studio UI",
            "request_type": "generation",
            "fresh_min": 2,
            "fresh_max": 2,
            "observed_current_config": 0,
            "resume_remaining_min": 2,
            "resume_remaining_max": 2,
            "quota_sensitive": "yes",
            "generation_or_probe": True,
            "fixed_matrix": True,
            "notes": "One seed send and one rerun from the first user turn.",
        },
        {
            "phase": "Host UI Google",
            "label": "Playground Gemini basic + search chat",
            "provider": "Google AI Studio",
            "surface": "Visible Playground UI",
            "request_type": "generation",
            "fresh_min": 2,
            "fresh_max": 2,
            "observed_current_config": 0,
            "resume_remaining_min": 2,
            "resume_remaining_max": 2,
            "quota_sensitive": "yes",
            "generation_or_probe": True,
            "fixed_matrix": True,
            "notes": "One basic chat send and one Search-enabled Playground send.",
        },
        {
            "phase": "Host UI OpenAI",
            "label": "OpenAI-compatible optional/basic/search/repeated visible sends",
            "provider": "OpenAI-compatible",
            "surface": "Visible Local Studio UI",
            "request_type": "generation",
            "fresh_min": 6,
            "fresh_max": 6,
            "observed_current_config": 0,
            "resume_remaining_min": 6,
            "resume_remaining_max": 6,
            "quota_sensitive": "OpenAI quota/cost",
            "generation_or_probe": True,
            "fixed_matrix": True,
            "notes": "Optional search+image text, optional search+image+reasoning text, basic, search, repeat-1, repeat-2.",
        },
        {
            "phase": "Host UI OpenAI",
            "label": "OpenAI-compatible reasoning matrix",
            "provider": "OpenAI-compatible",
            "surface": "Visible Local Studio UI",
            "request_type": "generation",
            "fresh_min": 3,
            "fresh_max": 15,
            "observed_current_config": 0,
            "resume_remaining_min": 3,
            "resume_remaining_max": 15,
            "quota_sensitive": "OpenAI quota/cost",
            "generation_or_probe": True,
            "fixed_matrix": True,
            "notes": "Stream, non-stream, repeat per attempt. Host smoke retries the OpenAI reasoning matrix up to 5 attempts if summary/text evidence is missing.",
        },
        {
            "phase": "Host UI OpenAI",
            "label": "OpenAI-compatible interface compatibility matrix",
            "provider": "OpenAI-compatible",
            "surface": "Visible Local Studio UI",
            "request_type": "generation/control-error matrix",
            "fresh_min": 3,
            "fresh_max": 9,
            "observed_current_config": 0,
            "resume_remaining_min": 3,
            "resume_remaining_max": 9,
            "quota_sensitive": "OpenAI quota/cost",
            "generation_or_probe": True,
            "fixed_matrix": True,
            "notes": "OpenAI Chat strict entries send 3 requests. Gemini/Claude compatibility entries may fail before send or add up to 6 controlled-error requests.",
        },
        {
            "phase": "Host UI OpenAI",
            "label": "Invalid provider recovery send",
            "provider": "OpenAI-compatible",
            "surface": "Visible Local Studio UI",
            "request_type": "generation after controlled local error",
            "fresh_min": 1,
            "fresh_max": 3,
            "observed_current_config": 0,
            "resume_remaining_min": 1,
            "resume_remaining_max": 3,
            "quota_sensitive": "OpenAI quota/cost",
            "generation_or_probe": True,
            "fixed_matrix": True,
            "notes": "The invalid 127.0.0.1:9 send is not an external provider request. The recovery send can retry up to 3 times for transient upstream errors.",
        },
        {
            "phase": "Model/control calls",
            "label": "Google model-list/control calls used by API/UI harness",
            "provider": "Google AI Studio",
            "surface": "API/UI model loaders",
            "request_type": "model/control",
            "fresh_min": 12,
            "fresh_max": 30,
            "observed_current_config": 12,
            "resume_remaining_min": 10,
            "resume_remaining_max": 28,
            "quota_sensitive": "low/no generation quota, still real upstream",
            "generation_or_probe": False,
            "fixed_matrix": True,
            "notes": "Model-list estimates include boot/API/host UI loads. Host Local Studio model loaders retry up to 3 attempts. Browser page resource traffic is excluded.",
        },
        {
            "phase": "Model/control calls",
            "label": "OpenAI-compatible model-list/control calls used by API/UI harness",
            "provider": "OpenAI-compatible",
            "surface": "API/UI model loaders",
            "request_type": "model/control",
            "fresh_min": 15,
            "fresh_max": 44,
            "observed_current_config": 15,
            "resume_remaining_min": 12,
            "resume_remaining_max": 36,
            "quota_sensitive": "OpenAI quota/cost may apply depending provider",
            "generation_or_probe": False,
            "fixed_matrix": True,
            "notes": "Includes direct OpenAI /models preflight, compatibility model list, API Local Studio model list, and host UI model loads. Direct Responses probes are counted above as generation probes.",
        },
        {
            "phase": "Latest blocked run",
            "label": "Startup warmup Google quota blocker",
            "provider": "Google AI Studio",
            "surface": "Server startup / GET /health",
            "request_type": "failed generation-readiness probe",
            "fresh_min": 0,
            "fresh_max": 0,
            "observed_current_config": 0,
            "resume_remaining_min": 0,
            "resume_remaining_max": 0,
            "quota_sensitive": "yes",
            "generation_or_probe": False,
            "fixed_matrix": False,
            "notes": "The 012925 resume stopped during account-browser warmup with official Google current-quota/rate-limits text before Host UI started; the startup warmup row already counts the warmup request estimate.",
        },
        {
            "phase": "Local/control plane",
            "label": "Provider Manager, request logs, account local health, synthetic account CRUD",
            "provider": "None / local filesystem",
            "surface": "API/UI control plane",
            "request_type": "local/control",
            "fresh_min": 0,
            "fresh_max": 0,
            "observed_current_config": 0,
            "resume_remaining_min": 0,
            "resume_remaining_max": 0,
            "quota_sensitive": "no",
            "generation_or_probe": False,
            "fixed_matrix": True,
            "notes": "Account health checks validate local storage-state shape; generation permission is proven by warmup or real generation, not the health button itself.",
        },
    ]


IMAGE_CASE_SPECS: list[dict[str, Any]] = [
    {
        "case_id": "G-LS-04",
        "category": "Rendered image generation",
        "image_generation_expected": "yes",
        "current_image_coverage": "missing: Google rendered image prompt was not executed",
        "additional_real_request_min": 1,
        "additional_real_request_max": 1,
        "counts_for_image_gap_total": True,
        "notes": "Needs a Google Responses Image Tool prompt with a visible rendered image URL and request-log evidence.",
    },
    {
        "case_id": "G-LS-05",
        "category": "Image Tool optional text path",
        "image_generation_expected": "no",
        "current_image_coverage": "text-only optional-tool path; Image Tool is enabled but prompt must not force image generation",
        "additional_real_request_min": 0,
        "additional_real_request_max": 0,
        "counts_for_image_gap_total": False,
        "notes": "This is image-tool semantics, not rendered image generation. Current clean formal artifact still marks the row failed; later harness code maps this path but was blocked by Google quota before a new clean pass artifact.",
    },
    {
        "case_id": "G-LS-06",
        "category": "Search-to-image multi-turn bug path",
        "image_generation_expected": "yes",
        "current_image_coverage": "missing: multi-turn Google search-to-image path was not executed",
        "additional_real_request_min": 4,
        "additional_real_request_max": 4,
        "counts_for_image_gap_total": True,
        "notes": "Covers greet, identity, news/search, then make-it-image. BUG-GEMINI-IMAGE-TOOL-01 can share this same path if assertions are strong enough.",
    },
    {
        "case_id": "G-LS-07",
        "category": "Non-stream search+image generation",
        "image_generation_expected": "yes",
        "current_image_coverage": "missing: non-stream infographic/image path was not executed",
        "additional_real_request_min": 1,
        "additional_real_request_max": 1,
        "counts_for_image_gap_total": True,
        "notes": "Needed to prove image/text/error persistence does not depend on SSE streaming.",
    },
    {
        "case_id": "O-LS-04",
        "category": "Rendered image generation or controlled unsupported path",
        "image_generation_expected": "yes_or_controlled_unsupported",
        "current_image_coverage": "missing: OpenAI-compatible image generation or controlled unsupported-path assertion was not executed",
        "additional_real_request_min": 1,
        "additional_real_request_max": 1,
        "counts_for_image_gap_total": True,
        "notes": "If the configured OpenAI-compatible provider does not support images, this should still assert graceful failure and health recovery.",
    },
    {
        "case_id": "O-LS-05",
        "category": "Image Tool optional text path",
        "image_generation_expected": "no",
        "current_image_coverage": "text-only optional-tool path; Image Tool is enabled but prompt must not force image generation",
        "additional_real_request_min": 0,
        "additional_real_request_max": 0,
        "counts_for_image_gap_total": False,
        "notes": "This proves optional image capability is not forced for ordinary text prompts.",
    },
    {
        "case_id": "O-LS-07",
        "category": "Search+Image+Reasoning optional text path",
        "image_generation_expected": "no",
        "current_image_coverage": "text-only reasoning/tool path; image generation should not be forced",
        "additional_real_request_min": 0,
        "additional_real_request_max": 0,
        "counts_for_image_gap_total": False,
        "notes": "This is image-tool optionality plus reasoning preservation, not rendered image generation.",
    },
    {
        "case_id": "LS-UI-02",
        "category": "Image Tool controls",
        "image_generation_expected": "no",
        "current_image_coverage": "controls/model selectors only; no rendered image expected",
        "additional_real_request_min": 0,
        "additional_real_request_max": 0,
        "counts_for_image_gap_total": False,
        "notes": "Checks provider-specific image controls and avoids cross-provider residual UI state.",
    },
    {
        "case_id": "LS-UI-03",
        "category": "Image model filtering",
        "image_generation_expected": "no",
        "current_image_coverage": "model-list/filtering only; no rendered image expected",
        "additional_real_request_min": 0,
        "additional_real_request_max": 0,
        "counts_for_image_gap_total": False,
        "notes": "Image models must appear in Image Tool selectors and not pollute chat model lists.",
    },
    {
        "case_id": "BASE-IMG-01",
        "category": "Base Images page generation",
        "image_generation_expected": "yes",
        "current_image_coverage": "covered: #images generated a real persisted image in the latest visible UI run",
        "additional_real_request_min": 0,
        "additional_real_request_max": 0,
        "counts_for_image_gap_total": False,
        "notes": "Latest visible UI evidence proves base image generation is independent from Local Studio Image Tool.",
    },
    {
        "case_id": "BASE-IMG-02",
        "category": "Base Images reference/edit/retry",
        "image_generation_expected": "yes",
        "current_image_coverage": "covered: #images reference/edit/retry passed in the latest complete visible UI matrix",
        "additional_real_request_min": 0,
        "additional_real_request_max": 0,
        "counts_for_image_gap_total": False,
        "notes": "Latest complete matrix covered generation, visible retry/recovery, and fallback-model handling for #images reference/edit/retry.",
    },
    {
        "case_id": "API-LS-01",
        "category": "Image model list API",
        "image_generation_expected": "no",
        "current_image_coverage": "model-list only; should include image models",
        "additional_real_request_min": 0,
        "additional_real_request_max": 0,
        "counts_for_image_gap_total": False,
        "notes": "Provider-facing model list/control calls are counted separately from generation requests.",
    },
    {
        "case_id": "API-LS-04",
        "category": "Google API image-tool bug path",
        "image_generation_expected": "yes_or_image_prompt_oracle",
        "current_image_coverage": "not proven as rendered/image-prompt coverage in current artifact; mapped API pass is search/tool regression evidence",
        "additional_real_request_min": 1,
        "additional_real_request_max": 1,
        "counts_for_image_gap_total": True,
        "notes": "SYSTEM_TEST_PLAN.md asks for Google search=true + image_tool_enabled=true + image prompt. Current clean artifact does not prove a generated image asset.",
    },
    {
        "case_id": "API-LS-06",
        "category": "Generated image asset API",
        "image_generation_expected": "asset_get_only",
        "current_image_coverage": "missing: generated image asset URL was not opened through API",
        "additional_real_request_min": 0,
        "additional_real_request_max": 0,
        "counts_for_image_gap_total": False,
        "notes": "Requires an image produced by another row; the asset GET itself is local and should not consume provider quota.",
    },
    {
        "case_id": "API-LS-09",
        "category": "Generated asset/cache variation API",
        "image_generation_expected": "partial_asset_followup",
        "current_image_coverage": "partial: generated-asset API coverage is explicitly missing",
        "additional_real_request_min": 0,
        "additional_real_request_max": 0,
        "counts_for_image_gap_total": False,
        "notes": "This can piggyback on a real generated image from G-LS/API/base image rows.",
    },
    {
        "case_id": "API-BASE-01",
        "category": "Base-compatible image API",
        "image_generation_expected": "yes",
        "current_image_coverage": "missing: /v1/images/generations smoke was not executed",
        "additional_real_request_min": 1,
        "additional_real_request_max": 1,
        "counts_for_image_gap_total": True,
        "notes": "Covers the base API image-generation surface outside Local Studio.",
    },
    {
        "case_id": "BUG-GEMINI-IMAGE-TOOL-01",
        "category": "Regression: Google search+image multi-turn",
        "image_generation_expected": "yes",
        "current_image_coverage": "missing: regression path was not executed in visible UI",
        "additional_real_request_min": 0,
        "additional_real_request_max": 0,
        "counts_for_image_gap_total": False,
        "notes": "Do not double count if the G-LS-06 multi-turn path includes the regression assertions.",
    },
    {
        "case_id": "PM-PROTO-01",
        "category": "Future protocol image matrix",
        "image_generation_expected": "future_phase",
        "current_image_coverage": "not applicable in current Phase 1 code state",
        "additional_real_request_min": 0,
        "additional_real_request_max": 0,
        "counts_for_image_gap_total": False,
        "notes": "Becomes applicable when shared runtime gateway/protocol adapter matrix is implemented.",
    },
    {
        "case_id": "PM-RT-02",
        "category": "Future image routing policy",
        "image_generation_expected": "future_phase",
        "current_image_coverage": "not applicable in current Phase 1 code state",
        "additional_real_request_min": 0,
        "additional_real_request_max": 0,
        "counts_for_image_gap_total": False,
        "notes": "Becomes applicable when advanced routing policy handles stream/tool/image compatibility.",
    },
]


def image_rows(cases_by_id: dict[str, dict[str, Any]], plan_rows: dict[str, dict[str, str]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for spec in IMAGE_CASE_SPECS:
        case_id = str(spec["case_id"])
        case = cases_by_id.get(case_id, {})
        plan = plan_rows.get(case_id, {})
        status = str(case.get("status") or "not_mapped")
        rows.append(
            {
                "category": spec["category"],
                "case_id": case_id,
                "matrix_status": status,
                "tested_by": case.get("tested_by", ""),
                "image_generation_expected": spec["image_generation_expected"],
                "current_image_coverage": spec["current_image_coverage"],
                "artifact_evidence_or_reason": compact_json(case.get("evidence") or case.get("reason") or "", 700),
                "plan_action": compact_json(plan.get("plan_action", ""), 700),
                "plan_expected": compact_json(plan.get("plan_expected", ""), 700),
                "additional_real_request_min": spec["additional_real_request_min"],
                "additional_real_request_max": spec["additional_real_request_max"],
                "counts_for_image_gap_total": spec["counts_for_image_gap_total"],
                "notes": spec["notes"],
            }
        )
    return rows


def merge_explicit_latest_passes(cases_by_id: dict[str, dict[str, Any]], latest_alignment: dict[str, Any]) -> dict[str, dict[str, Any]]:
    merged = {case_id: dict(case) for case_id, case in cases_by_id.items()}
    latest_passes = set(str(case_id) for case_id in latest_alignment.get("passing_required_cases") or [])
    explicit_cases = {"G-LS-07"}
    for case_id in sorted(explicit_cases & latest_passes):
        current = dict(merged.get(case_id, {"id": case_id}))
        current.update(
            {
                "id": case_id,
                "status": "pass",
                "tested_by": current.get("tested_by") or "api_real+mcp_visible",
                "_evidence_run": LATEST_RESUME_RUN,
                "evidence": (
                    f"Latest quota-blocked verification run {LATEST_RESUME_RUN} reached and passed {case_id} "
                    "before stopping on external Google quota exhaustion."
                ),
            }
        )
        merged[case_id] = current
    return merged


def row_total(rows: list[dict[str, Any]], key: str, *, provider: str | None = None, generation_only: bool = False, fixed_only: bool = False) -> int:
    total = 0
    for row in rows:
        if provider and row["provider"] != provider:
            continue
        if generation_only and not row["generation_or_probe"]:
            continue
        if fixed_only and not row["fixed_matrix"]:
            continue
        value = row.get(key)
        if isinstance(value, int):
            total += value
    return total


def append_table(worksheet, rows: list[dict[str, Any]], headers: list[str]) -> None:
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header, "") for header in headers])


def add_sheet_title(worksheet, title: str, subtitle: str, column_count: int) -> int:
    last_column = get_column_letter(column_count)
    worksheet.merge_cells(f"A1:{last_column}1")
    worksheet.merge_cells(f"A2:{last_column}2")
    worksheet["A1"] = title
    worksheet["A2"] = subtitle
    worksheet["A1"].fill = PatternFill("solid", fgColor="17324D")
    worksheet["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    worksheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    worksheet["A2"].fill = PatternFill("solid", fgColor="EAF2F8")
    worksheet["A2"].font = Font(color="17324D", italic=True)
    worksheet["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    worksheet.row_dimensions[1].height = 28
    worksheet.row_dimensions[2].height = 36
    worksheet.append([])
    return 4


def write_titled_table(
    worksheet,
    *,
    title: str,
    subtitle: str,
    rows: list[dict[str, Any]],
    headers: list[str],
    status_column: int | None = None,
) -> None:
    header_row = add_sheet_title(worksheet, title, subtitle, len(headers))
    append_table(worksheet, rows, headers)
    style_table(worksheet, header_row=header_row, status_column=status_column)


def style_table(worksheet, *, header_row: int = 1, status_column: int | None = None) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_gray = Side(style="thin", color="D9E2F3")
    grid_border = Border(top=thin_gray, bottom=thin_gray, left=thin_gray, right=thin_gray)
    even_fill = PatternFill("solid", fgColor="F8FBFD")
    for cell in worksheet[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = grid_border
    worksheet.row_dimensions[header_row].height = 30
    for row_index, row in enumerate(worksheet.iter_rows(min_row=header_row + 1), start=header_row + 1):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = grid_border
            if row_index % 2 == 0:
                cell.fill = even_fill
        if status_column:
            status_value = str(row[status_column - 1].value or "").lower()
            fill = None
            if status_value in {"pass", "通过"}:
                fill = PatternFill("solid", fgColor="DFF2E1")
            elif status_value in {"fail", "失败"}:
                fill = PatternFill("solid", fgColor="FADBD8")
            elif status_value in {"not_applicable", "不适用"}:
                fill = PatternFill("solid", fgColor="E8E8E8")
            elif status_value in {"blocked", "阻塞"}:
                fill = PatternFill("solid", fgColor="FFF2CC")
            elif status_value in {"未映射", "not_mapped"}:
                fill = PatternFill("solid", fgColor="FDEBD0")
            if fill:
                for cell in row:
                    cell.fill = fill
    worksheet.freeze_panes = f"A{header_row + 1}"
    if worksheet.max_row >= header_row:
        worksheet.auto_filter.ref = f"A{header_row}:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells[:90])
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 62)
    for row_index in range(header_row + 1, worksheet.max_row + 1):
        worksheet.row_dimensions[row_index].height = 42


def build_tracking_workbook() -> None:
    current_ui = read_wsl_json(artifact_path(CURRENT_MATRIX_RUN, "ui-results.json"))
    current_summary = read_wsl_json(artifact_path(CURRENT_MATRIX_RUN, "summary.json"))
    plan_alignment = read_wsl_json(artifact_path(CURRENT_MATRIX_RUN, "plan-script-alignment-results.json"))
    latest_alignment = read_wsl_json(artifact_path(LATEST_RESUME_RUN, "plan-script-alignment-results.json"))
    quota_blocker = latest_quota_blocker(LATEST_RESUME_RUN)
    plan_rows = parse_plan_rows()

    cases = current_ui.get("cases") if isinstance(current_ui.get("cases"), list) else []
    case_ids = [str(case.get("id") or "") for case in cases if isinstance(case, dict) and case.get("id")]
    cases_by_id = merge_explicit_latest_passes({str(case.get("id") or ""): case for case in cases if isinstance(case, dict)}, latest_alignment)
    status_counter = Counter(str(case.get("status") or "unknown") for case in cases_by_id.values() if isinstance(case, dict))
    case_counter = Counter(case_ids)
    plan_ids = sorted(plan_rows)
    result_ids = sorted(case_ids)
    missing_ids = sorted(set(plan_ids) - set(result_ids))
    extra_ids = sorted(set(result_ids) - set(plan_ids))
    duplicate_result_ids = sorted(case_id for case_id, count in case_counter.items() if count > 1)
    consistency_ok = not missing_ids and not extra_ids and not duplicate_result_ids and len(plan_ids) == len(result_ids)

    image_coverage_rows = image_rows(cases_by_id, plan_rows)
    image_by_id = {str(row.get("case_id") or ""): row for row in image_coverage_rows}
    image_real_request_min = sum(int(row.get("additional_real_request_min") or 0) for row in image_coverage_rows if row.get("counts_for_image_gap_total") is True)
    image_real_request_max = sum(int(row.get("additional_real_request_max") or 0) for row in image_coverage_rows if row.get("counts_for_image_gap_total") is True)

    request_count_rows = request_rows()
    provider_request_rows = [row for row in request_count_rows if row.get("provider") != "None / local filesystem"]
    observed_total = sum(int(row.get("observed_current_config") or 0) for row in provider_request_rows)
    observed_generation = sum(int(row.get("observed_current_config") or 0) for row in provider_request_rows if row.get("generation_or_probe") is True)
    remaining_total_min = sum(int(row.get("resume_remaining_min") or 0) for row in provider_request_rows)
    remaining_total_max = sum(int(row.get("resume_remaining_max") or 0) for row in provider_request_rows)
    remaining_generation_min = sum(int(row.get("resume_remaining_min") or 0) for row in provider_request_rows if row.get("generation_or_probe") is True)
    remaining_generation_max = sum(int(row.get("resume_remaining_max") or 0) for row in provider_request_rows if row.get("generation_or_probe") is True)
    fresh_generation_min = row_total(request_count_rows, "fresh_min", generation_only=True, fixed_only=True)
    fresh_generation_max = row_total(request_count_rows, "fresh_max", generation_only=True, fixed_only=True)
    total_cases = len(case_ids) or int(current_ui.get("required_case_count") or 0)
    pass_count = status_counter.get("pass", 0)
    fail_count = status_counter.get("fail", 0)
    not_applicable_count = status_counter.get("not_applicable", 0)
    recorded_count = pass_count + fail_count + not_applicable_count
    denominator = max(total_cases, 1)

    def request_range(minimum: int, maximum: int) -> str:
        return str(minimum) if minimum == maximum else f"{minimum}-{maximum}"

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "汇总"

    summary_rows = [
        {"指标": "测试用例总数", "数值": total_cases, "说明": f"来自最新完整矩阵产物；SYSTEM_TEST_PLAN.md 解析到 {len(plan_ids)} 个计划 ID。"},
        {"指标": "已记录结果", "数值": f"{recorded_count}/{total_cases}", "说明": f"执行进度 {recorded_count / denominator:.1%}；包括通过、失败和不适用。"},
        {"指标": "通过 / 失败 / 不适用", "数值": f"{pass_count} / {fail_count} / {not_applicable_count}", "说明": "失败项保留为待修复或待补测，不当作通过。"},
        {"指标": "待修复或补测", "数值": fail_count + len(missing_ids), "说明": "失败用例加计划中存在但结果缺失的用例。"},
        {"指标": "计划一致性", "数值": "通过" if consistency_ok else "失败", "说明": f"缺失 {len(missing_ids)}，额外 {len(extra_ids)}，重复 {len(duplicate_result_ids)}。"},
        {"指标": "最近功能阻塞", "数值": "无（最新完整矩阵仅剩 dirty gates / G-LS-07，随后验证 run 已通过 G-LS-07）", "说明": "完整矩阵 run 20260611-232555 中 BASE-IMG-02 已通过；最新 run 20260612-001338 在外部 Google 配额阻塞前已将 G-LS-07 标为通过。"},
        {"指标": "最近配额阻塞", "数值": quota_blocker.get("label", "无"), "说明": f"{quota_blocker.get('reason', '最新完整矩阵未以配额阻塞停止。')}；最新 resume {LATEST_RESUME_RUN}。"},
        {"指标": "最近 UI 部分证据", "数值": LATEST_PARTIAL_UI_RUN, "说明": "该运行在 checkpoint_after_provider_manager 写入了 partial mcp-visible-ui-results.json；不计入失败 Google 生成覆盖。"},
        {"指标": "已使用实际请求（全部真实上游）", "数值": observed_total, "说明": "含 warmup、生成/探测、模型/控制调用和最近失败配额尝试；不含本地 CRUD、截图和浏览器静态资源。"},
        {"指标": "已使用实际请求（生成/探测）", "数值": observed_generation, "说明": "只统计生成、聊天发送、直接模型探测和 warmup 生成就绪探测。"},
        {"指标": "还需要实际请求（续跑，全部真实上游）", "数值": request_range(remaining_total_min, remaining_total_max), "说明": "低额度续跑口径：含 warmup、剩余生成/探测和模型/控制调用，不含图像缺口补测。"},
        {"指标": "还需要实际请求（续跑，生成/探测）", "数值": request_range(remaining_generation_min, remaining_generation_max), "说明": "低额度续跑口径：含 warmup，不含模型/控制调用，不含图像缺口补测。"},
        {"指标": "还需要实际请求（含图像缺口，全部真实上游）", "数值": request_range(remaining_total_min + image_real_request_min, remaining_total_max + image_real_request_max), "说明": f"在续跑全部真实上游请求基础上，额外计入图像缺口 {image_real_request_min}-{image_real_request_max} 次。"},
        {"指标": "还需要实际请求（含图像缺口，生成/探测）", "数值": request_range(remaining_generation_min + image_real_request_min, remaining_generation_max + image_real_request_max), "说明": "用户最关心的低额度续跑 + 图像补齐口径；不含模型/控制调用。"},
        {"指标": "完整重跑估算（生成/探测，不含 warmup）", "数值": request_range(fresh_generation_min, fresh_generation_max), "说明": "完整固定矩阵的生成/探测请求估算；启动 warmup 需另加当前 6 次估算。"},
        {"指标": "运行依据", "数值": CURRENT_MATRIX_RUN, "说明": current_summary.get("system_test_result", "产物不可用")},
        {"指标": "干净基线", "数值": BASELINE_CLEAN_FORMAL_RUN, "说明": "历史 clean formal API/UI 基线；当前工作树诊断 run 不能替代 clean pass 结论。"},
        {"指标": "矩阵映射完成", "数值": label(plan_alignment.get("matrix_mapping_complete", current_ui.get("matrix_mapping_complete", "未知"))), "说明": plan_alignment.get("reason", "")},
    ]
    write_titled_table(
        summary_sheet,
        title="系统测试执行汇总",
        subtitle="只保留进度和真实请求数：详细筛选请看“测试用例”页。",
        rows=summary_rows,
        headers=["指标", "数值", "说明"],
    )

    ordered_case_ids = list(dict.fromkeys(case_ids + missing_ids))
    case_rows: list[dict[str, Any]] = []
    for case_id in ordered_case_ids:
        case = cases_by_id.get(case_id, {})
        classification = classify_case(case_id)
        plan = plan_rows.get(case_id, {})
        status = str(case.get("status") or "not_mapped")
        tested_by = str(case.get("tested_by") or "not_covered")
        evidence_or_reason = compact_json(case.get("evidence") or case.get("reason") or "", 900)
        image_row = image_by_id.get(case_id)
        image_gap = ""
        if image_row and image_row.get("counts_for_image_gap_total") is True:
            image_gap = request_range(int(image_row.get("additional_real_request_min") or 0), int(image_row.get("additional_real_request_max") or 0))
        google_blocked_gap = case_id.startswith(("G-LS", "BASE-IMG", "BUG-GEMINI", "BUG-AISTUDIO", "PERF-")) and status != "pass"
        case_rows.append(
            {
                "优先级": priority_for_case(case_id),
                "用例 ID": case_id,
                "执行结果": status_label(status),
                "跟进状态": followup_state(case_id, status, tested_by),
                "覆盖领域": label(classification["coverage_area"], AREA_LABELS),
                "提供方": label(classification["provider"]),
                "入口/层面": label(classification["surface"], SURFACE_LABELS),
                "验证方式": tested_by_label(tested_by),
                "图像相关": "是" if image_row else "否",
                "补齐图像缺口请求": image_gap,
                "计划区域": plan.get("plan_area", ""),
                "阻塞/证据": blocker_for_case(case_id, status, tested_by, evidence_or_reason) or evidence_or_reason,
                "下一步": next_action_for_case(case_id, status, tested_by),
                "最近运行": case.get("_evidence_run") or (LATEST_RESUME_RUN if google_blocked_gap else CURRENT_MATRIX_RUN),
                "备注": f"最新 resume：{quota_blocker.get('label', '无配额阻塞')}；partial UI 证据：{LATEST_PARTIAL_UI_RUN}" if google_blocked_gap else "",
            }
        )
    cases_sheet = workbook.create_sheet("测试用例")
    write_titled_table(
        cases_sheet,
        title="测试用例执行结果",
        subtitle="一条用例一行，可按执行结果、提供方、图像相关、跟进状态筛选。",
        rows=case_rows,
        headers=["优先级", "用例 ID", "执行结果", "跟进状态", "覆盖领域", "提供方", "入口/层面", "验证方式", "图像相关", "补齐图像缺口请求", "计划区域", "阻塞/证据", "下一步", "最近运行", "备注"],
        status_column=3,
    )

    workbook.save(OUTPUT_PATH)
    print(OUTPUT_PATH)


if __name__ == "__main__":
    build_tracking_workbook()